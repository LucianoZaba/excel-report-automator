import pandas as pd
from pathlib import Path
from openpyxl.styles import Font
from .pipeline import ReadError, CleanError, ExportError, log

def read(path: str) -> pd.DataFrame:
    try:
        log.info(f"Leyendo {path}")
        p = Path(path)
        if p.suffix.lower() == '.csv':
            return pd.read_csv(path)
        return pd.read_excel(path)
    except Exception as e:
        raise ReadError(f"No se pudo leer {path}: {e}")

def clean(df: pd.DataFrame) -> pd.DataFrame:
    try:
        original_len = len(df)
        df = df.copy()

        # 1. Normaliza columnas
        df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
        # Fix: si hay columnas duplicadas despues de normalizar (ej ' Cliente ' y 'cliente'), quedarse con la primera
        df = df.loc[:, ~df.columns.duplicated(keep='first')]

        required = ['cliente', 'producto', 'cantidad']
        has_precio = 'precio' in df.columns or 'precio_unitario' in df.columns
        missing = [c for c in required if c not in df.columns]
        if missing or not has_precio:
            if not has_precio:
                missing.append('precio / precio_unitario')
            raise CleanError(f"Faltan columnas: {missing}")

        # 2. Limpieza strings generica
        for col in df.select_dtypes(include=['object']).columns:
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace({'nan': pd.NA, 'None': pd.NA, '': pd.NA, 'NAT': pd.NA})

        df = df.dropna(how='all')

        # 2.1 Normalizacion especifica de texto clave (case-insensitive)
        # Cliente, vendedor, sucursal -> Title Case para evitar SOFIA vs Sofia
        for col in ['cliente', 'vendedor', 'sucursal', 'producto']:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].str.title()
                df[col] = df[col].replace({'Nan': pd.NA})

        # 3. Tipos numericos
        for col in ['cantidad', 'precio', 'precio_unitario', 'descuento', 'id']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        # 4. Fechas
        if 'fecha' in df.columns:
            df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce')
            invalid = df['fecha'].isna().sum()
            if invalid > 0:
                log.warning(f"Fechas invalidas eliminadas: {invalid}")
                df = df.dropna(subset=['fecha'])

        # 5. Validaciones
        if 'cantidad' in df.columns:
            before = len(df)
            df = df[df['cantidad'] > 0]
            if len(df) < before:
                log.warning(f"Cantidad <=0 eliminadas: {before - len(df)}")

        precio_col = 'precio_unitario' if 'precio_unitario' in df.columns else 'precio'
        if precio_col in df.columns:
            before = len(df)
            df = df[(df[precio_col] >= 0) & (df[precio_col] <= 10_000_000)]
            if len(df) < before:
                log.warning(f"Precios imposibles eliminados: {before - len(df)}")

        if 'descuento' in df.columns:
            max_desc = df['descuento'].max()
            if pd.notna(max_desc) and max_desc > 1:
                df['descuento'] = df['descuento'] / 100.0
            before = len(df)
            df = df[(df['descuento'] >= 0) & (df['descuento'] <= 1)]
            if len(df) < before:
                log.warning(f"Descuentos fuera de 0-100% eliminados: {before - len(df)}")
            df['descuento'] = df['descuento'].fillna(0)
        else:
            df['descuento'] = 0.0

        # 6. Estado -> ENTREGADO / PENDIENTE / CANCELADO
        if 'estado' in df.columns:
            df['estado'] = df['estado'].astype(str).str.strip().str.upper()

        # 7. CUIT: 1 cliente = 1 CUIT (case-insensitive)
        if 'cliente' in df.columns and 'cuit' in df.columns:
            # agrupar por cliente en minuscula para detectar SOFIA DIAZ vs Sofia Diaz
            df['_cliente_norm'] = df['cliente'].astype(str).str.lower().str.strip()
            multi = df.groupby('_cliente_norm')['cuit'].nunique()
            conflictivos = multi[multi > 1]
            if not conflictivos.empty:
                for norm in conflictivos.index:
                    sub = df[df['_cliente_norm'] == norm]
                    cuits = sub['cuit'].unique()
                    nombres = sub['cliente'].unique()
                    log.warning(f"CUIT inconsistente para cliente_norm='{norm}' nombres={list(nombres)} cuits={list(cuits)[:3]}... -> unificado al mas frecuente")
            # unificar CUIT al mas frecuente por cliente_norm
            cuit_mode = df.groupby('_cliente_norm')['cuit'].agg(lambda x: x.mode().iloc[0] if not x.mode().empty else x.iloc[0])
            df['cuit'] = df['_cliente_norm'].map(cuit_mode)
            # opcional: unificar nombre cliente al mas frecuente
            cliente_mode = df.groupby('_cliente_norm')['cliente'].agg(lambda x: x.mode().iloc[0] if not x.mode().empty else x.iloc[0])
            df['cliente'] = df['_cliente_norm'].map(cliente_mode)
            df = df.drop(columns=['_cliente_norm'])

        # 8. IDs unicos
        if 'id' in df.columns:
            before = len(df)
            df = df.drop_duplicates(subset=['id'], keep='first')
            if len(df) < before:
                log.warning(f"IDs duplicados eliminados: {before - len(df)}")
            if df['id'].duplicated().any():
                raise CleanError("IDs duplicados persisten")

        # 9. Duplicados exactos
        df = df.drop_duplicates()

        # 10. Total con descuento
        if 'cantidad' in df.columns and precio_col in df.columns:
            df['total'] = df['cantidad'] * df[precio_col] * (1 - df['descuento'])

        log.info(f"Limpieza: {original_len} -> {len(df)} filas")
        return df.reset_index(drop=True)

    except CleanError:
        raise
    except Exception as e:
        raise CleanError(f"Error limpiando: {e}")

def export(df: pd.DataFrame, output_path: str = "data/processed/reporte_limpio.xlsx") -> Path:
    try:
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        df.to_excel(out, index=False, sheet_name="Reporte")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(out)
            ws = wb["Reporte"]
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.auto_filter.ref = ws.dimensions
            for col in ws.columns:
                max_len = max(len(str(c.value)) if c.value else 0 for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)
            ws.freeze_panes = "A2"
            wb.save(out)
        except Exception as fmt_e:
            log.warning(f"Formato openpyxl: {fmt_e}")
        return out
    except Exception as e:
        raise ExportError(f"Error exportando: {e}")
